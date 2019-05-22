import time
from dataclasses import dataclass
from typing import List
from openpyxl import load_workbook
from helpers import get_vendor_from_id, get_product_with_product_type, get_delivery_from_del_number
from input_data.load_customers import Customer
from input_data.load_vendors import load_vendors, Vendor
from input_data.products import ProductSpec
from optimize.optimize import start_optimize, Action
from profit import calculate_profit
from scenarios.load_scenarios import Scenario
from simulation.stochastic import optimize_with_one_product_type_at_the_time
from solution_method import SolutionMethod


@dataclass(frozen=True)
class ResultFromOneRun:
    running_time: float
    actions: List[Action]
    profit: float


def run_simulation(
        customers: List[Customer],
        number_of_runs_in_one_scenario: int,
        product_specs: List[ProductSpec],
        scenarios: List[Scenario],
        start_day: int,
        number_of_days_in_each_run: int,
        solution_method: SolutionMethod,
        one_product_type_at_the_time: bool,
        adjust_delivery_estimate: float,
        end_day: float,
):
    profit_for_scenarios = []
    average_time_for_scenarios = []

    sheet, workbook = _open_excel()

    next_empty_row_in_excel_after_intro = _print_input_to_excel(adjust_delivery_estimate, end_day, number_of_days_in_each_run, one_product_type_at_the_time,
                          sheet, start_day, solution_method.name)

    for scenario_index, scenario in enumerate(scenarios):
        number_of_columns_in_each_scenario = 7
        next_empty_row_in_excel = next_empty_row_in_excel_after_intro
        next_empty_row_in_excel = _print_scenario_header_to_excel(next_empty_row_in_excel, scenario_index, sheet, number_of_columns_in_each_scenario)

        print("Simulating over scenario " + str(scenario_index))
        vendors = load_vendors(
            path="input_data/deliveries.xlsx",
            adjust_delivery_estimate=adjust_delivery_estimate,
        )
        current_start_day = start_day
        profits_for_scenario = []
        running_times_for_scenario = []

        for run_number in range(number_of_runs_in_one_scenario):
            print("--------------------------")
            print("Run number " + str(run_number) + " in scenario " + str(scenario_index))
            current_end_day = current_start_day + number_of_days_in_each_run - 1
            print("Optimize from day " + str(current_start_day) + " to day " + str(current_end_day))

            sheet.cell(row=next_empty_row_in_excel, column=1 + number_of_columns_in_each_scenario * scenario_index).value = run_number

            update_todays_deliveries_based_on_actual_volume_in_scenario(
                vendors=vendors,
                scenario=scenario,
                current_day=current_start_day,
            )
            vendors_with_relevant_deliveries_for_next_time_period = _filter_out_deliveries_after_end_time(
                vendors=vendors,
                end_day=current_end_day,
            )
            customers_with_relevant_orders_for_next_time_period = _filter_out_order_out_of_time_scope(
                customers=customers,
                start_day=current_start_day,
                end_day=current_end_day,
            )
            number_of_relevant_orders = len([
                order
                for customer in customers_with_relevant_orders_for_next_time_period
                for order in customer.orders
                if order.departure_day == current_start_day
            ])
            print("number of orders with delivery today: " + str(number_of_relevant_orders))
            if number_of_relevant_orders > 0:
                start_time = time.time()
                if one_product_type_at_the_time:
                    actions, objective_value = optimize_with_one_product_type_at_the_time(
                        vendors=vendors_with_relevant_deliveries_for_next_time_period,
                        customers=customers_with_relevant_orders_for_next_time_period,
                        product_specs=product_specs,
                        solution_method=solution_method,
                        number_of_days_in_each_run=number_of_days_in_each_run,
                        start_day=start_day,
                    )
                else:
                    actions, objective_value = start_optimize(
                        vendors=vendors_with_relevant_deliveries_for_next_time_period,
                        customers=customers_with_relevant_orders_for_next_time_period,
                        product_specs=product_specs,
                        solution_method=solution_method,
                        number_of_days_in_each_run=number_of_days_in_each_run,
                        start_day=start_day,
                        include_cross_docking=True,
                    )
                end_time = time.time()
                total_time = end_time - start_time
                print("Run time: " + str(total_time) + " sec")
                running_times_for_scenario.append(total_time)
            else:
                actions = []

            actions_with_transportation_date_today = [
                action
                for action in actions
                if action.transportation_day == current_start_day
            ]
            profit_from_todays_operation = calculate_profit(
                vendors=vendors,
                customers=customers,
                product_specs=product_specs,
                todays_actions=actions_with_transportation_date_today,
            )
            _print_run_results_to_excel(actions_with_transportation_date_today, current_start_day,
                                        customers_with_relevant_orders_for_next_time_period, next_empty_row_in_excel,
                                        number_of_columns_in_each_scenario, profit_from_todays_operation,
                                        scenario_index, sheet, vendors_with_relevant_deliveries_for_next_time_period)
            next_empty_row_in_excel += 1

            print("Profit from day " + str(current_start_day) + " = " + str(profit_from_todays_operation))
            profits_for_scenario.append(profit_from_todays_operation)

            update_delivery_volumes_after_todays_operations(
                vendors=vendors,
                todays_actions=actions_with_transportation_date_today,
            )
            current_start_day += 1

        total_profit_for_scenario = sum(profits_for_scenario)
        average_time_for_scenario = sum(running_times_for_scenario) / len(running_times_for_scenario)
        print("profit for scenario " + str(scenario_index) + ": " + str(total_profit_for_scenario))

        profit_for_scenarios.append(total_profit_for_scenario)
        average_time_for_scenarios.append(average_time_for_scenario)

        sheet.cell(row=next_empty_row_in_excel,
                   column=1 + number_of_columns_in_each_scenario * scenario_index).value = "Total profit"
        sheet.cell(row=next_empty_row_in_excel,
                   column=2 + number_of_columns_in_each_scenario * scenario_index).value = total_profit_for_scenario
        next_empty_row_in_excel += 1
        sheet.cell(row=next_empty_row_in_excel,
                   column=1 + number_of_columns_in_each_scenario * scenario_index).value = "Average time"
        sheet.cell(row=next_empty_row_in_excel,
                   column=2 + number_of_columns_in_each_scenario * scenario_index).value = average_time_for_scenario
        next_empty_row_in_excel += 1

    average_run_time = sum(average_time_for_scenarios) / len(average_time_for_scenarios)
    print("Average run time: " + str(average_run_time))
    average_profit = sum(profit_for_scenarios) / len(profit_for_scenarios)
    print("Average profit is: " + str(average_profit))

    next_empty_row_in_excel += 5

    sheet.cell(row=next_empty_row_in_excel, column=1).value = "Average profit"
    sheet.cell(row=next_empty_row_in_excel, column=2).value = average_profit
    next_empty_row_in_excel += 1
    sheet.cell(row=next_empty_row_in_excel, column=1).value = "Average run time"
    sheet.cell(row=next_empty_row_in_excel, column=2).value = average_run_time

    workbook.save("simulation/PythonEksport.xlsx")


def _print_run_results_to_excel(actions_with_transportation_date_today, current_start_day,
                                customers_with_relevant_orders_for_next_time_period, next_empty_row_in_excel,
                                number_of_columns_in_each_scenario, profit_from_todays_operation, scenario_index, sheet,
                                vendors_with_relevant_deliveries_for_next_time_period):
    x_value = sum([
        action.volume_delivered
        for action in actions_with_transportation_date_today
        if action.internal_delivery
    ])
    y_value = sum([
        action.volume_delivered
        for action in actions_with_transportation_date_today
        if not action.internal_delivery
    ])
    supply = sum([
        product_supply.volume
        for vendor in vendors_with_relevant_deliveries_for_next_time_period
        for delivery in vendor.deliveries
        if delivery.arrival_day <= current_start_day
        for product_supply in delivery.supply
    ])
    demand = sum([
        product_demand.volume
        for customer in customers_with_relevant_orders_for_next_time_period
        for order in customer.orders
        if order.departure_day == current_start_day
        for product_demand in order.demand
    ])
    sheet.cell(row=next_empty_row_in_excel,
               column=2 + number_of_columns_in_each_scenario * scenario_index).value = profit_from_todays_operation
    sheet.cell(row=next_empty_row_in_excel,
               column=3 + number_of_columns_in_each_scenario * scenario_index).value = x_value
    sheet.cell(row=next_empty_row_in_excel,
               column=4 + number_of_columns_in_each_scenario * scenario_index).value = y_value
    sheet.cell(row=next_empty_row_in_excel,
               column=5 + number_of_columns_in_each_scenario * scenario_index).value = supply
    sheet.cell(row=next_empty_row_in_excel,
               column=6 + number_of_columns_in_each_scenario * scenario_index).value = demand


def _print_scenario_header_to_excel(next_empty_row_in_excel, scenario_index, sheet, number_of_columns_in_each_scenario):
    sheet.cell(row=next_empty_row_in_excel,
               column=1 + number_of_columns_in_each_scenario * scenario_index).value = "Scenario"
    sheet.cell(row=next_empty_row_in_excel,
               column=2 + number_of_columns_in_each_scenario * scenario_index).value = scenario_index
    next_empty_row_in_excel += 1
    sheet.cell(row=next_empty_row_in_excel,
               column=1 + number_of_columns_in_each_scenario * scenario_index).value = "Day"
    sheet.cell(row=next_empty_row_in_excel,
               column=2 + number_of_columns_in_each_scenario * scenario_index).value = "Profit"
    sheet.cell(row=next_empty_row_in_excel,
               column=3 + number_of_columns_in_each_scenario * scenario_index).value = "Internal X"
    sheet.cell(row=next_empty_row_in_excel,
               column=4 + number_of_columns_in_each_scenario * scenario_index).value = "External Y"
    sheet.cell(row=next_empty_row_in_excel,
               column=5 + number_of_columns_in_each_scenario * scenario_index).value = "Supply"
    sheet.cell(row=next_empty_row_in_excel,
               column=6 + number_of_columns_in_each_scenario * scenario_index).value = "Demand"
    next_empty_row_in_excel += 1
    return next_empty_row_in_excel


def _print_input_to_excel(adjust_delivery_estimate, end_day, number_of_days_in_each_run, one_product_type_at_the_time,
                          sheet, start_day, stochastic):
    sheet.cell(row=1, column=1).value = "Days in a run"
    sheet.cell(row=1, column=2).value = number_of_days_in_each_run
    sheet.cell(row=2, column=1).value = "Start day"
    sheet.cell(row=2, column=2).value = start_day
    sheet.cell(row=3, column=1).value = "End day"
    sheet.cell(row=3, column=2).value = end_day
    sheet.cell(row=4, column=1).value = "Stochastic"
    sheet.cell(row=4, column=2).value = stochastic
    sheet.cell(row=5, column=1).value = "One product at a time"
    sheet.cell(row=5, column=2).value = one_product_type_at_the_time
    sheet.cell(row=6, column=1).value = "adjust delivery estimate"
    sheet.cell(row=6, column=2).value = adjust_delivery_estimate

    return 8


def _open_excel():
    workbook = load_workbook("simulation/PythonEksport.xlsx")
    workbook.create_sheet("New sheet")
    sheets = workbook.sheetnames
    number_of_sheets = len(sheets)
    sheet = workbook[sheets[number_of_sheets - 1]]
    return sheet, workbook


def _filter_out_deliveries_after_end_time(vendors: List[Vendor], end_day: int) -> List[Vendor]:
    vendors_with_only_relevant_deliveries = [
        Vendor(
            id=vendor.id,
            transportation_cost_per_box=vendor.transportation_cost_per_box,
            deliveries=[
                delivery
                for delivery in vendor.deliveries
                if delivery.arrival_day <= end_day
            ]
        )
        for vendor in vendors
    ]
    return vendors_with_only_relevant_deliveries


def _filter_out_order_out_of_time_scope(customers: List[Customer], start_day: int, end_day: int) -> List[Customer]:
    customer_with_only_relevant_deliveries = [
        Customer(
            id=customer.id,
            orders=[
                order
                for order in customer.orders
                if start_day <= order.departure_day <= end_day
            ],
            customer_category=customer.customer_category,
            out_of_country=customer.out_of_country,
            transportation_price_per_box=customer.transportation_price_per_box
        )
        for customer in customers
    ]
    return customer_with_only_relevant_deliveries


def update_todays_deliveries_based_on_actual_volume_in_scenario(vendors: List[Vendor], scenario: Scenario, current_day: int):
    for product_outcome in scenario.product_outcomes:
        vendor = get_vendor_from_id(vendors=vendors, vendor_id=product_outcome.vendor_id)
        delivery = get_delivery_from_del_number(deliveries=vendor.deliveries, delivery_number=product_outcome.delivery_number)
        if delivery.arrival_day == current_day:
            product_type = product_outcome.product_type
            product = get_product_with_product_type(products=delivery.supply, product_type=product_type)
            product.volume = product_outcome.actual_volume


def update_delivery_volumes_after_todays_operations(vendors: List[Vendor], todays_actions: List[Action]):
    for action in todays_actions:
        if action.internal_delivery:
            vendor = get_vendor_from_id(vendors, action.vendor_id)
            delivery = get_delivery_from_del_number(deliveries=vendor.deliveries, delivery_number=action.delivery_number)

            product = get_product_with_product_type(products=delivery.supply, product_type=action.product_type)
            product.volume -= action.volume_delivered

